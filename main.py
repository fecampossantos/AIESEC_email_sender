#-*- coding:utf-8 -*-

import properties

import csv

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import smtplib, ssl

from openpyxl import Workbook, load_workbook

#content = {'receiver_name': nome da pessoa que vai receber o sugar,
#           'receiver_email': email da pessoa,
#           'message': mensagem enviada,
#           'sender': jeito que a pessoa se declarou, anônimo/a se vazio}

def create_email(content):

    message = MIMEMultipart("alternative")
    message["Subject"] = "Você recebeu um sugar!"
    message["From"] = properties.sender
    message["To"] = content['receiver_email']

    # Turn these into plain/html MIMEText objects
    email_as_text = MIMEText(properties.getText(content['message'], content['sender']), "plain")

    email_as_html = MIMEText(properties.getHTML(content['message'], content['sender']), "html")

    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(email_as_text)
    message.attach(email_as_html)

    return message


def main():
    workbookName = ""
    sheetName = ""

    wb = load_workbook(filename="{}.xlsx".format(workbookName), data_only=True)
    sheet = wb[sheetName]

    for message, sender, receiver, email in sheet.iter_rows(min_row=2, min_col=1, max_col=4):
        if email.value is None:
            break
        content = {'receiver_name': receiver.value,
                   'receiver_email': email.value,
                   'message': message.value,
                   'sender': sender.value
                   }

        print(f"Sending email to {receiver.value}")
        
        message = create_email(content)

        context = ssl.create_default_context()

        with smtplib.SMTP_SSL("smtp.gmail.com", properties.port, context=context) as server:
            # logs to the server as sender
            server.login(properties.sender, properties.password)

            # sends the email
            server.sendmail(
                properties.sender, content['receiver_email'],
                message.as_string()
            )

        print(f"Email to {receiver.value} sent!")



if __name__ == '__main__':
    main()